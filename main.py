from owlready2 import *
from openpyxl import Workbook


def render_using_name(entity):
     return entity.name


set_render_func(render_using_name)

onto = get_ontology(r"C:\Users\smarotta\Documents\personal_onto.owl").load()

classes = list(onto.classes())
properties = list(onto.object_properties())

if __name__ == "__main__":
    wb = Workbook()
    classes_ws = wb.create_sheet("classes")
    properties_ws = wb.create_sheet("properties")

    line = 2
    for c in classes:
        if c.comment:
            desc = c.comment[0].replace("\n", " ")
        else:
            desc = ""
        sources = list()
        if c.source:
            for s in c.source:
                sources.append(s)
        if c.seeAlso:
            for s in c.seeAlso:
                sources.append(s)
        parents = map(str, list(c.is_a))
        subclasses = map(str, list(c.subclasses()))

        classes_ws[f"A{line}"] = "Class name"
        classes_ws[f"A{line + 1}"] = "Class description"
        classes_ws[f"A{line + 2}"] = "Class parents"
        classes_ws[f"A{line + 3}"] = "Subclasses"
        classes_ws[f"A{line + 4}"] = "Sources"

        classes_ws[f"B{line}"] = str(c)
        classes_ws[f"B{line + 1}"] = desc
        classes_ws[f"B{line + 2}"] = ", ".join(parents)
        classes_ws[f"B{line + 3}"] = ", ".join(subclasses)
        classes_ws[f"B{line + 4}"] = ", ".join(sources)
        line += 6

    line = 2
    for p in properties:
        name = p.name
        desc = p.comment[0]
        domain = list(map(render_using_name, list(p.domain)))
        range = list(map(render_using_name, list(p.range)))

        properties_ws[f"A{line}"] = "Property name"
        properties_ws[f"A{line + 1}"] = "Property description"
        properties_ws[f"A{line + 2}"] = "Property domain"
        properties_ws[f"A{line + 3}"] = "Property range"

        properties_ws[f"B{line}"] = name
        properties_ws[f"B{line + 1}"] = desc
        properties_ws[f"B{line + 2}"] = ", ".join(domain)
        properties_ws[f"B{line + 3}"] = ", ".join(range)
        line += 5

    wb.save("onto.xlsx")


# adding comments to properties automatically
# for p in properties:
#     print(p.name)
#     if p.comment:
#         comment = f"original: {p.comment[0]}"
#     else:
#         domain = p.domain
#         if len(domain) == 1:
#             comment = f"Links {p.domain[0]} to {p.range[0]}".replace("_", " ")
#         else:
#             comment = f"Links some entites to {p.range[0]}".replace("_", " ")
#         p.comment = comment
#     print(comment)
# onto.save(file = "onto.owl", format = "rdfxml")